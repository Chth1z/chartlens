from __future__ import annotations

import json
from io import BytesIO
from typing import Any

from openpyxl import Workbook

from app.core.config_loader import load_export_template
from app.domain.models import ExportTemplate, ValidatedFieldResult


def build_export_workbook(results: list[ValidatedFieldResult], template: ExportTemplate | None = None) -> bytes:
    template = template or load_export_template()
    by_key = {result.field_key: result for result in results}
    wb = Workbook()
    ws = wb.active
    ws.title = "EYEX"
    audit = wb.create_sheet("Evidence Audit")

    for col_index, column in enumerate(template.columns, start=1):
        ws.cell(row=1, column=col_index, value=column.header)
        result = by_key.get(column.field_key)
        ws.cell(row=2, column=col_index, value=_export_value(result, column.unknown_value, template))

    audit.append(
        [
            "field_key",
            "value",
            "status",
            "confidence",
            "auto_accepted",
            "validation_state",
            "risk_level",
            "review_required",
            "exportable",
            "export_gate_reason",
            "provenance",
            "acceptance_reason",
            "model_profile_id",
            "ocr_engine",
            "evidence_pack_hash",
            "token_estimate",
            "llm_cache_status",
            "ocr_page_quality",
            "evidence_span",
            "evidence_block_id",
            "page",
            "bbox",
            "error_code",
            "validator_messages",
            "reasoning",
        ]
    )
    for result in results:
        audit.append(
            [
                result.field_key,
                result.normalized_code,
                result.status,
                result.confidence,
                result.auto_accepted,
                result.validation_state,
                result.risk_level,
                result.review_required,
                _is_exportable(result, template),
                _export_gate_reason(result, template),
                _excel_value(result.provenance),
                result.acceptance_reason,
                result.model_profile_id,
                result.ocr_engine,
                _primary_pack_hash(result),
                _primary_token_estimate(result),
                result.provenance.get("llm_cache_status") or result.provenance.get("cache_status"),
                _excel_value(result.provenance.get("ocr_page_quality")),
                result.evidence_span,
                result.evidence_block_id,
                result.page,
                _excel_value(result.bbox),
                result.error_code,
                "；".join(result.validator_messages),
                result.reasoning_summary,
            ]
        )

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def _excel_value(value: Any) -> Any:
    if isinstance(value, (dict, list, tuple)):
        return json.dumps(value, ensure_ascii=False, separators=(",", ":"))
    return value


def _export_value(result: ValidatedFieldResult | None, column_unknown: str | None, template: ExportTemplate) -> str:
    if not _is_exportable(result, template):
        return template.unknown_value if column_unknown is None else column_unknown
    return result.normalized_code


def _is_exportable(result: ValidatedFieldResult | None, template: ExportTemplate) -> bool:
    if result is None or result.review_required or result.normalized_code in (None, "unknown"):
        return False
    if not template.export_gate.require_pass_or_reviewed:
        return True
    if (
        result.validation_state in template.export_gate.reviewed_states
        or result.acceptance_reason in template.export_gate.manual_acceptance_reasons
    ):
        return True
    return (
        result.validation_state == "accepted"
        and result.provenance.get("decision_status") == template.export_gate.pass_decision_status
    )


def _export_gate_reason(result: ValidatedFieldResult, template: ExportTemplate) -> str:
    if result.review_required:
        return "review_required"
    if result.normalized_code in (None, "unknown"):
        return "unknown_or_missing_value"
    if not template.export_gate.require_pass_or_reviewed:
        return "legacy_gate_allowed"
    if (
        result.validation_state in template.export_gate.reviewed_states
        or result.acceptance_reason in template.export_gate.manual_acceptance_reasons
    ):
        return "manual_review"
    if (
        result.validation_state == "accepted"
        and result.provenance.get("decision_status") == template.export_gate.pass_decision_status
    ):
        return "pass_decision"
    if (
        result.provenance.get("decision_status")
        and result.provenance.get("decision_status") != template.export_gate.pass_decision_status
    ):
        return "decision_not_pass"
    return "not_pass_or_reviewed"


def _primary_pack_hash(result: ValidatedFieldResult) -> str | None:
    if result.evidence_packs:
        return result.evidence_packs[0].pack_hash
    if result.evidence_candidates:
        return result.evidence_candidates[0].pack_hash
    return None


def _primary_token_estimate(result: ValidatedFieldResult) -> int:
    if result.evidence_packs:
        return result.evidence_packs[0].token_estimate
    if result.evidence_candidates:
        return result.evidence_candidates[0].token_estimate
    return 0
