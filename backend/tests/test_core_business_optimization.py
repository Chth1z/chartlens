import json
from io import BytesIO
from types import SimpleNamespace

from fastapi.testclient import TestClient
from openpyxl import load_workbook
from sqlalchemy import select

from app.core.config_loader import load_document_profile, load_extraction_schema
from app.core.database import CaseRecord, FieldResultRecord, ReviewAuditRecord, SessionLocal
from app.core.settings import settings
from app.domain.models import DocumentIR, DocumentIRBlock, ExtractionCandidate, ValidatedFieldResult
from app.main import app
from app.services import ocr
from app.services.deidentify import deidentify_document_ir
from app.services.pipeline import process_case
from app.services.llm_provider.fallback import ModelFallbackProvider
from app.services.llm_provider.types import SemanticExtractionProvider
from app.services.export import build_export_workbook
from app.services.validation import validate_candidate


class _AlwaysFailingProvider(SemanticExtractionProvider):
    name = "always-failing"
    route = "online_llm"

    def extract_group(self, *, document_ir, group, fields, blocks):
        raise RuntimeError("upstream unavailable")

    def collect_evidence(self, *, document_context, fields):
        raise RuntimeError("upstream unavailable")


def _profile():
    return SimpleNamespace(
        model_ref="deepseek/deepseek-v4-flash",
        provider_id="deepseek",
        profile_id="deepseek_v4_flash",
        model="deepseek-v4-flash",
        provider="openai_compatible",
    )


def test_model_fallback_returns_unknown_for_complex_fields_after_all_models_fail(monkeypatch):
    schema = load_extraction_schema()
    field = schema.field_by_key("hypertension_history")
    group = schema.group_by_key(field.field_group_key)
    block = DocumentIRBlock(
        block_id="b1",
        page=1,
        reading_order=1,
        text="既往史：否认高血压。",
        section_label="既往史",
        confidence=0.98,
    )
    monkeypatch.setattr("app.services.llm_provider.fallback._provider_for_profile", lambda profile: _AlwaysFailingProvider())

    fallback = ModelFallbackProvider([_profile()])
    candidates = fallback.extract_group(
        document_ir=DocumentIR(document_id="case-1", profile_id="medical_inpatient_zh", source_filename="case.txt", blocks=[block]),
        group=group,
        fields=[field],
        blocks=[block],
    )

    assert fallback.route == "unknown_after_model_fallback"
    assert candidates[0].field_key == "hypertension_history"
    assert candidates[0].normalized_code == "unknown"
    assert candidates[0].evidence_type == "no_evidence"
    assert candidates[0].review_required is True
    assert candidates[0].error_code == "LLM_PROVIDER_FAILED"
    assert fallback.last_usage["fallback_errors"] == [
        "deepseek/deepseek-v4-flash: RuntimeError: upstream unavailable"
    ]


def test_review_rejects_nonexistent_evidence_span():
    db = SessionLocal()
    case_id = "CASE-OPT-REVIEW"
    try:
        db.query(FieldResultRecord).filter(FieldResultRecord.case_id == case_id).delete()
        db.query(CaseRecord).filter(CaseRecord.case_id == case_id).delete()
        db.commit()
        document_ir = DocumentIR(
            document_id=case_id,
            profile_id="medical_inpatient_zh",
            source_filename="case.txt",
            blocks=[
                DocumentIRBlock(
                    block_id="b1",
                    page=1,
                    reading_order=1,
                    text="既往史：否认高血压。",
                    section_label="既往史",
                    confidence=0.98,
                )
            ],
        )
        db.add(
            CaseRecord(
                case_id=case_id,
                filename="case.txt",
                file_hash="hash",
                file_path="case.txt",
                status="completed",
                document_ir_json=document_ir.model_dump_json(),
            )
        )
        db.add(
            FieldResultRecord(
                case_id=case_id,
                field_key="hypertension_history",
                payload_json=ValidatedFieldResult(
                    field_key="hypertension_history",
                    field_group_key="history_group",
                    normalized_code="unknown",
                    status="unknown",
                    review_required=True,
                ).model_dump_json(),
            )
        )
        db.commit()

        response = TestClient(app).post(
            f"/api/cases/{case_id}/review",
            json={
                "field_key": "hypertension_history",
                "normalized_code": "0",
                "raw_value": "无",
                "evidence_span": "糖尿病",
                "evidence_block_id": "b1",
            },
        )
    finally:
        db.query(FieldResultRecord).filter(FieldResultRecord.case_id == case_id).delete()
        db.query(CaseRecord).filter(CaseRecord.case_id == case_id).delete()
        db.commit()
        db.close()

    assert response.status_code == 400
    assert "evidence_span" in response.json()["detail"]


def test_manual_review_without_evidence_is_audited_and_exported():
    db = SessionLocal()
    case_id = "CASE-OPT-MANUAL-REVIEW"
    try:
        db.query(ReviewAuditRecord).filter(ReviewAuditRecord.case_id == case_id).delete()
        db.query(FieldResultRecord).filter(FieldResultRecord.case_id == case_id).delete()
        db.query(CaseRecord).filter(CaseRecord.case_id == case_id).delete()
        db.commit()
        db.add(
            CaseRecord(
                case_id=case_id,
                filename="case.txt",
                file_hash="hash",
                file_path="case.txt",
                status="completed",
                document_ir_json=DocumentIR(
                    document_id=case_id,
                    profile_id="medical_inpatient_zh",
                    source_filename="case.txt",
                    blocks=[
                        DocumentIRBlock(
                            block_id="b1",
                            page=1,
                            reading_order=1,
                            text="入院记录未明确HH分级。",
                            section_label="入院记录",
                            confidence=0.98,
                        )
                    ],
                ).model_dump_json(),
            )
        )
        db.add(
            FieldResultRecord(
                case_id=case_id,
                field_key="hh_grade",
                payload_json=ValidatedFieldResult(
                    field_key="hh_grade",
                    field_group_key="score_group",
                    normalized_code="unknown",
                    status="unknown",
                    review_required=True,
                    evidence_type="no_evidence",
                    acceptance_reason="unknown_or_insufficient_evidence",
                ).model_dump_json(),
            )
        )
        db.commit()
        client = TestClient(app)

        response = client.post(
            f"/api/cases/{case_id}/review",
            json={
                "field_key": "hh_grade",
                "normalized_code": "2",
                "raw_value": "2",
                "reviewer": "local-reviewer",
                "comment": "人工复核确认",
            },
        )

        assert response.status_code == 200
        payload = response.json()
        assert payload["normalized_code"] == "2"
        assert payload["review_required"] is False
        assert payload["validation_state"] == "reviewed"
        assert payload["evidence_span"] is None
        assert payload["provenance"]["manual_review_without_document_evidence"] is True

        audits = db.execute(
            select(ReviewAuditRecord).where(ReviewAuditRecord.case_id == case_id)
        ).scalars().all()
        assert len(audits) == 1
        assert json.loads(audits[0].after_json)["normalized_code"] == "2"

        summary = client.get(f"/api/cases/{case_id}").json()
        assert summary["audit_count"] == 1

        export_response = client.get(f"/api/cases/{case_id}/export")
        assert export_response.status_code == 200
        workbook = load_workbook(BytesIO(export_response.content), data_only=True)
        sheet = workbook["EYEX"]
        headers = [cell.value for cell in sheet[1]]
        assert sheet.cell(row=2, column=headers.index("HH分组") + 1).value == "2"
    finally:
        db.query(ReviewAuditRecord).filter(ReviewAuditRecord.case_id == case_id).delete()
        db.query(FieldResultRecord).filter(FieldResultRecord.case_id == case_id).delete()
        db.query(CaseRecord).filter(CaseRecord.case_id == case_id).delete()
        db.commit()
        db.close()


def test_export_serializes_structured_audit_values():
    data = build_export_workbook(
        [
            ValidatedFieldResult(
                field_key="hypertension_history",
                field_group_key="history_group",
                normalized_code="0",
                status="confirmed",
                review_required=False,
                provenance={
                    "ocr_page_quality": {
                        "page": 1,
                        "quality_band": "good",
                    }
                },
            )
        ]
    )

    workbook = load_workbook(BytesIO(data), data_only=True)
    audit = workbook["Evidence Audit"]
    headers = [cell.value for cell in audit[1]]
    value = audit.cell(row=2, column=headers.index("ocr_page_quality") + 1).value
    assert value == '{"page":1,"quality_band":"good"}'


def test_export_only_writes_pass_or_reviewed_values():
    data = build_export_workbook(
        [
            ValidatedFieldResult(
                field_key="gender",
                field_group_key="demographics_group",
                normalized_code="1",
                status="confirmed",
                confidence=0.99,
                review_required=False,
                validation_state="accepted",
                provenance={"decision_status": "REVIEW"},
                acceptance_reason="model_confidence_without_pass",
            ),
            ValidatedFieldResult(
                field_key="age",
                field_group_key="demographics_group",
                normalized_code="16",
                status="confirmed",
                confidence=0.99,
                review_required=False,
                validation_state="accepted",
                provenance={"decision_status": "PASS"},
                acceptance_reason="explicit_evidence",
            ),
            ValidatedFieldResult(
                field_key="hospital",
                field_group_key="demographics_group",
                normalized_code="河北医科大学第四医院",
                status="confirmed",
                confidence=0.7,
                review_required=False,
                validation_state="reviewed",
                provenance={"manual_review_without_document_evidence": True},
                acceptance_reason="manual_review",
            ),
        ]
    )

    workbook = load_workbook(BytesIO(data), data_only=True)
    sheet = workbook["EYEX"]
    headers = [cell.value for cell in sheet[1]]

    assert sheet.cell(row=2, column=headers.index("性别(男1，女2)") + 1).value == "unknown"
    assert sheet.cell(row=2, column=headers.index("年龄") + 1).value == "16"
    assert sheet.cell(row=2, column=headers.index("医院") + 1).value == "河北医科大学第四医院"

    audit = workbook["Evidence Audit"]
    audit_headers = [cell.value for cell in audit[1]]
    assert audit.cell(row=2, column=audit_headers.index("exportable") + 1).value is False
    assert audit.cell(row=2, column=audit_headers.index("export_gate_reason") + 1).value == "decision_not_pass"
    assert audit.cell(row=3, column=audit_headers.index("export_gate_reason") + 1).value == "pass_decision"
    assert audit.cell(row=4, column=audit_headers.index("export_gate_reason") + 1).value == "manual_review"


def test_deidentification_blocks_online_when_residual_phi_remains():
    document_ir = DocumentIR(
        document_id="case-risk",
        profile_id="medical_inpatient_zh",
        source_filename="case.txt",
        metadata={
            "ocr_engine_candidates": [
                {
                    "engine": "fixture",
                    "alternative_blocks": [
                        {"text": "姓名：张三，电话：13800138000，现住深圳市南山区科技园", "confidence": 0.8}
                    ],
                }
            ]
        },
        blocks=[
            DocumentIRBlock(
                block_id="b1",
                page=1,
                reading_order=1,
                text="现病史：患者居住在深圳市南山区科技园，否认高血压。",
                section_label="现病史",
                confidence=0.98,
            )
        ],
    )

    redacted = deidentify_document_ir(document_ir, load_document_profile())

    deidentification = redacted.metadata["deidentification"]
    assert deidentification["risk_score"] > 0
    assert deidentification["online_llm_allowed"] is False
    assert "RESIDUAL_ADDRESS_PATTERN" in deidentification["risk_findings"]
    candidate_text = redacted.metadata["ocr_engine_candidates"][0]["alternative_blocks"][0]["text"]
    assert "张三" not in candidate_text
    assert "13800138000" not in candidate_text


def test_image_document_ir_uses_cache_and_adds_source_metadata(monkeypatch, tmp_path):
    image_path = tmp_path / "case.png"
    image_path.write_bytes(b"not-a-real-image")
    calls = {"count": 0}

    def fake_intelligent_extract(file_path, aliases):
        calls["count"] += 1
        return [
            DocumentIRBlock(
                block_id="b0001-fixture",
                page=1,
                reading_order=1,
                text="姓名：张三",
                confidence=0.95,
                block_type="form_field",
            )
        ], {
            "ocr_adapter": "intelligent_document",
            "ocr_engine": "fixture_intelligent",
            "ocr_intelligent_status": "completed",
        }

    monkeypatch.setattr(settings, "storage_dir", tmp_path / "storage")
    monkeypatch.setattr(ocr, "extract_with_intelligent_ocr", fake_intelligent_extract)

    first = ocr.build_document_ir(image_path, image_path.read_bytes(), document_id="case-cache-1")
    second = ocr.build_document_ir(image_path, image_path.read_bytes(), document_id="case-cache-2")

    assert calls["count"] == 1
    assert first.metadata["ocr_cache_status"] == "miss"
    assert second.metadata["ocr_cache_status"] == "hit"
    assert second.blocks[0].source_engine == "fixture_intelligent"
    assert second.blocks[0].source_page_kind == "image_ocr"
    assert second.blocks[0].ocr_profile == settings.ocr_profile
    assert "low_confidence" not in second.blocks[0].quality_flags


def test_complex_fact_then_code_field_without_facts_requires_review():
    schema = load_extraction_schema()
    field = schema.field_by_key("aneurysm_location")
    document_ir = DocumentIR(
        document_id="case-facts",
        profile_id="medical_inpatient_zh",
        source_filename="case.txt",
        blocks=[
            DocumentIRBlock(
                block_id="b1",
                page=1,
                reading_order=1,
                text="CTA示前交通动脉瘤。",
                section_label="辅助检查",
                confidence=0.98,
            )
        ],
    )
    candidate = ExtractionCandidate(
        field_key=field.key,
        field_group_key=field.field_group_key,
        normalized_code="3",
        evidence_span="前交通动脉瘤",
        evidence_block_id="b1",
        evidence_type="event_fact",
        confidence=0.92,
        review_required=False,
    )

    result = validate_candidate(candidate, field, document_ir)

    assert result.normalized_code == "3"
    assert result.review_required is True
    assert result.auto_accepted is False
    assert result.error_code == "COMPLEX_FIELD_REQUIRES_FACTS"
    assert result.validation_state == "needs_review"
    assert result.risk_level == "high"


def test_batch_eval_reports_business_quality_metrics():
    db = SessionLocal()
    case_id = "CASE-OPT-EVAL"
    try:
        db.query(FieldResultRecord).filter(FieldResultRecord.case_id == case_id).delete()
        db.query(CaseRecord).filter(CaseRecord.case_id == case_id).delete()
        db.commit()
        db.add(
            CaseRecord(
                case_id=case_id,
                filename="case.txt",
                file_hash="hash",
                file_path="case.txt",
                status="completed",
                diagnostics_json=json.dumps({"llm_usage": [{"usage": {"input_tokens": 10, "output_tokens": 3, "cost_usd": 0.01}}]}),
            )
        )
        db.add_all(
            [
                FieldResultRecord(
                    case_id=case_id,
                    field_key="hypertension_history",
                    payload_json=ValidatedFieldResult(
                        field_key="hypertension_history",
                        field_group_key="history_group",
                        normalized_code="0",
                        status="confirmed",
                        confidence=0.95,
                        evidence_span="否认高血压",
                        evidence_block_id="b1",
                        auto_accepted=True,
                        review_required=False,
                    ).model_dump_json(),
                ),
                FieldResultRecord(
                    case_id=case_id,
                    field_key="diabetes_history",
                    payload_json=ValidatedFieldResult(
                        field_key="diabetes_history",
                        field_group_key="history_group",
                        normalized_code="1",
                        status="confirmed",
                        confidence=0.91,
                        evidence_span="糖尿病",
                        evidence_block_id="b2",
                        auto_accepted=True,
                        review_required=False,
                    ).model_dump_json(),
                ),
            ]
        )
        db.commit()

        response = TestClient(app).post(
            "/api/evals/batch",
            json={"cases": [{"case_id": case_id, "gold": {"hypertension_history": "0", "diabetes_history": "unknown"}}]},
        )
    finally:
        db.query(FieldResultRecord).filter(FieldResultRecord.case_id == case_id).delete()
        db.query(CaseRecord).filter(CaseRecord.case_id == case_id).delete()
        db.commit()
        db.close()

    assert response.status_code == 200
    payload = response.json()
    assert payload["summary"]["total_fields"] == 2
    assert payload["summary"]["accuracy"] == 0.5
    assert payload["summary"]["auto_accept_precision"] == 0.5
    assert payload["summary"]["unknown_misfill_rate"] == 1.0
    assert payload["summary"]["evidence_coverage"] == 1.0


def test_processing_stores_raw_document_ir_as_protected_payload(monkeypatch, tmp_path):
    db = SessionLocal()
    case_id = "CASE-OPT-RAW"
    file_path = tmp_path / "case.txt"
    file_path.write_text("基本信息：患者，男，66岁。\n既往史：否认高血压。", encoding="utf-8")
    try:
        db.query(FieldResultRecord).filter(FieldResultRecord.case_id == case_id).delete()
        db.query(CaseRecord).filter(CaseRecord.case_id == case_id).delete()
        db.commit()
        case = CaseRecord(
            case_id=case_id,
            filename="case.txt",
            file_hash="hash",
            file_path=str(file_path),
            status="queued",
        )
        db.add(case)
        db.commit()
        db.refresh(case)
        monkeypatch.setattr(settings, "llm_mode", "disabled")

        process_case(db, case)

        db.refresh(case)
        protected = json.loads(case.raw_document_ir_json)
    finally:
        db.query(FieldResultRecord).filter(FieldResultRecord.case_id == case_id).delete()
        db.query(CaseRecord).filter(CaseRecord.case_id == case_id).delete()
        db.commit()
        db.close()

    assert protected["scheme"] == "win32-dpapi"
    assert "基本信息" not in protected["value"]
