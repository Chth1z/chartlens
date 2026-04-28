from io import BytesIO

from openpyxl import load_workbook

from app.schemas.pipeline import FieldExtractionResult
from app.services.exporter import build_excel_workbook
from app.services.field_dictionary import load_field_dictionary


def test_build_excel_workbook_uses_fixed_header_order_and_audit_sheet():
    dictionary = load_field_dictionary()
    results = [
        FieldExtractionResult(
            field_key="gender",
            raw_value="男",
            normalized_code="1",
            confidence=0.98,
            evidence_text="性别：男",
            page=1,
            bbox=[10, 10, 20, 20],
            reasoning_summary="字段直接出现。",
            review_required=False,
            error_code=None,
        ),
        FieldExtractionResult(
            field_key="hypertension_history",
            raw_value="有",
            normalized_code="1",
            confidence=0.88,
            evidence_text="高血压病史10年",
            page=1,
            bbox=[10, 40, 90, 55],
            reasoning_summary="既往史明确提及。",
            review_required=True,
            error_code=None,
        ),
    ]

    content = build_excel_workbook("CASE-001", dictionary, results)
    workbook = load_workbook(BytesIO(content))

    assert workbook.sheetnames == ["structured_data", "evidence_audit"]
    sheet = workbook["structured_data"]
    headers = [cell.value for cell in sheet[1]]
    values = [cell.value for cell in sheet[2]]

    assert headers[:4] == ["case_id", "性别(男1，女2)", "年龄", "医院"]
    assert values[0] == "CASE-001"
    assert values[1] == "1"
    audit = workbook["evidence_audit"]
    assert audit["A2"].value == "CASE-001"
    assert audit["B2"].value == "gender"
