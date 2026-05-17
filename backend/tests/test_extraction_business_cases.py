from app.core.config_loader import load_document_profile
from app.domain.models import DocumentIR, DocumentIRBlock
from app.services.deidentify import deidentify_document_ir
from app.services.export import build_export_workbook
from app.services.layout_normalizer import normalize_document_layout
from app.services.pipeline import extract_document
from app.services.llm_provider.local_extraction import ConservativeLocalProvider
from io import BytesIO
from openpyxl import load_workbook


def _ir(blocks: list[tuple[str, str]]) -> DocumentIR:
    return DocumentIR(
        document_id="case-fixture",
        profile_id="medical_inpatient_zh",
        source_filename="fixture.txt",
        blocks=[
            DocumentIRBlock(
                block_id=f"b{index}",
                page=1,
                reading_order=index,
                text=text,
                confidence=0.98,
                section_label=section,
            )
            for index, (section, text) in enumerate(blocks, start=1)
        ],
    )


def _extract(blocks: list[tuple[str, str]]):
    document_ir = deidentify_document_ir(_ir(blocks), load_document_profile())
    results = extract_document(document_ir, provider=ConservativeLocalProvider())
    return {result.field_key: result for result in results}


def _redacted_ir(blocks: list[tuple[str, str]]) -> DocumentIR:
    return deidentify_document_ir(_ir(blocks), load_document_profile())


def test_not_mentioned_is_unknown_not_negative():
    results = _extract(
        [
            ("既往史", "既往史：否认高血压，否认糖尿病。"),
            ("个人史", "个人史：无烟酒不良嗜好。"),
        ]
    )

    assert results["hypertension_history"].normalized_code == "0"
    assert results["diabetes_history"].normalized_code == "0"
    assert results["smoking_history"].normalized_code == "0"
    assert results["drinking_history"].normalized_code == "0"
    assert results["hyperlipidemia_history"].normalized_code == "unknown"


def test_family_history_does_not_mark_patient_stroke_positive():
    results = _extract(
        [
            ("家族史", "家族史：其父有脑梗死病史。"),
            ("既往史", "既往史：否认高血压。"),
        ]
    )

    assert results["stroke_history"].normalized_code == "unknown"
    assert results["stroke_history"].error_code in {"NOT_MENTIONED", "NON_PATIENT_EXPERIENCER"}


def test_uncertain_condition_requires_review_and_unknown():
    results = _extract([("既往史", "既往史：糖尿病？待排，否认高血压。")])

    assert results["diabetes_history"].normalized_code == "unknown"
    assert results["diabetes_history"].review_required is True


def test_aneurysm_and_surgery_use_fact_then_code():
    results = _extract(
        [
            ("辅助检查", "CTA示前交通动脉瘤，考虑责任动脉瘤，蛛网膜下腔出血。"),
            ("手术记录", "先行脑室外引流术，后行前交通动脉瘤开颅夹闭术。"),
        ]
    )

    assert results["aneurysm_location"].normalized_code == "3"
    assert results["aneurysm_location"].facts
    assert results["surgery_method"].normalized_code == "开颅夹闭术"
    assert results["surgery_method"].facts


def test_gcs_derived_wfns_is_review_candidate():
    results = _extract([("体格检查", "体格检查：GCS 10分，可见局灶神经功能缺损。")])

    assert results["wfns_grade"].normalized_code == "4"
    assert results["wfns_grade"].status == "derived_candidate"
    assert results["wfns_grade"].review_required is True


def test_urban_residence_uses_safe_pre_redaction_derivation():
    document_ir = _redacted_ir([("基本信息", "一般资料：患者现住深圳市南山区科技园，性别：男，年龄：58岁。")])
    joined_text = "\n".join(block.text for block in document_ir.blocks)

    assert "深圳市南山区科技园" not in joined_text
    assert "是否城市判定：城市" in joined_text
    assert document_ir.metadata["deidentification"]["online_llm_allowed"] is False

    results = extract_document(document_ir, provider=ConservativeLocalProvider())
    by_key = {result.field_key: result for result in results}

    assert by_key["urban_residence"].normalized_code == "2"
    assert by_key["urban_residence"].evidence_span == "是否城市判定：城市"
    assert "深圳市南山区科技园" not in (by_key["urban_residence"].evidence_text or "")


def test_timeline_fields_extract_explicit_chinese_durations():
    results = _extract(
        [
            ("现病史", "患者突发头痛3小时后入院，伴恶心呕吐。"),
            ("手术记录", "入院后2天行前交通动脉瘤开颅夹闭术。"),
        ]
    )

    assert results["onset_to_admission_time"].normalized_code == "3小时"
    assert results["onset_to_admission_time"].evidence_span == "突发头痛3小时后入院"
    assert results["admission_to_surgery_time"].normalized_code == "2天"
    assert results["admission_to_surgery_time"].evidence_span == "入院后2天行前交通动脉瘤开颅夹闭术"


def test_table_cell_demographics_flow_from_layout_to_export():
    profile = load_document_profile()
    raw = DocumentIR(
        document_id="case-table-cells",
        profile_id="medical_inpatient_zh",
        source_filename="case.pdf",
        blocks=[
            DocumentIRBlock(
                block_id="c-hospital",
                page=1,
                reading_order=1,
                text="河北医科大学第四医院",
                bbox=[180, 40, 620, 80],
                confidence=0.98,
                block_type="title",
                section_label="基本信息",
            ),
            DocumentIRBlock(
                block_id="c-gender-label",
                page=1,
                reading_order=2,
                text="性别",
                bbox=[120, 130, 180, 158],
                confidence=0.97,
                block_type="cell",
                section_label="基本信息",
                table_id="t-home",
                row=1,
                col=1,
            ),
            DocumentIRBlock(
                block_id="c-gender-value",
                page=1,
                reading_order=3,
                text="男",
                bbox=[190, 130, 230, 158],
                confidence=0.97,
                block_type="cell",
                section_label="基本信息",
                table_id="t-home",
                row=1,
                col=2,
            ),
            DocumentIRBlock(
                block_id="c-age-label",
                page=1,
                reading_order=4,
                text="年龄",
                bbox=[260, 130, 320, 158],
                confidence=0.97,
                block_type="cell",
                section_label="基本信息",
                table_id="t-home",
                row=1,
                col=3,
            ),
            DocumentIRBlock(
                block_id="c-age-value",
                page=1,
                reading_order=5,
                text="58",
                bbox=[330, 130, 380, 158],
                confidence=0.97,
                block_type="cell",
                section_label="基本信息",
                table_id="t-home",
                row=1,
                col=4,
            ),
            DocumentIRBlock(
                block_id="b-history",
                page=1,
                reading_order=6,
                text="既往史：否认高血压、糖尿病。",
                bbox=[100, 260, 720, 300],
                confidence=0.98,
                section_label="既往史",
            ),
        ],
    )

    document_ir = deidentify_document_ir(normalize_document_layout(raw, profile), profile)
    assert any(block.block_id == "c-gender-label" and block.block_type == "cell" for block in document_ir.blocks)
    assert any(block.block_id == "c-gender-value" and block.block_type == "cell" for block in document_ir.blocks)
    results = extract_document(document_ir, provider=ConservativeLocalProvider())
    by_key = {result.field_key: result for result in results}

    assert by_key["gender"].normalized_code == "1"
    assert by_key["gender"].evidence_span == "性别：男"
    assert by_key["gender"].evidence_block_id in {block.block_id for block in document_ir.blocks}
    assert by_key["age"].normalized_code == "58"
    assert by_key["age"].evidence_span == "年龄：58"
    assert by_key["hypertension_history"].normalized_code == "0"
    assert by_key["diabetes_history"].normalized_code == "0"

    workbook = load_workbook(BytesIO(build_export_workbook(results)), data_only=True)
    sheet = workbook["EYEX"]
    headers = [cell.value for cell in sheet[1]]
    assert sheet.cell(row=2, column=headers.index("性别(男1，女2)") + 1).value == "1"
    assert sheet.cell(row=2, column=headers.index("年龄") + 1).value == "58"
